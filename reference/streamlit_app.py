import copy
import datetime
import io
import re

import altair as alt
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mg_weather_openmeteo import getweatherdatamixed, resolvetimezoneandelevation
from mg_pv_core import (
    LOSSPARAMS,
    GeneratorConfig,
    computepvforgenerator,
    defaultlosssettings,
    defaultobstacles,
    lossparamsforgenerator,
)


st.set_page_config(
    page_title="PV-Forecast",
    layout="wide",
    page_icon="☀️",
    initial_sidebar_state="expanded",
)


FIELD_HELP_TEXTS = {
    "gamma": (
        "태양광 모듈 온도가 기준온도보다 올라갈 때 출력이 얼마나 감소하는지를 나타내는 값입니다.\n\n"
        "추천값:\n"
        "- 다결정 실리콘 모듈: -0.40 ~ -0.45 %/°C\n"
        "- 단결정 실리콘 모듈: -0.30 ~ -0.40 %/°C\n"
        "- 일반 결정질 실리콘 모듈: -0.35 %/°C\n"
        "- 화합물 반도체(CIGS/CdTe): -0.20 ~ -0.35 %/°C\n"
        "- 비정질 실리콘(a-Si): -0.15 ~ -0.25 %/°C"
    ),
    "bifaciality": (
        "양면형 모듈의 후면 발전 효율을 나타내는 값입니다. 70%는 후면으로 들어오는 "
        "반사광을 전면과 같은 효율로 100% 발전하는 것이 아니라, 전면 발전량 대비 약 "
        "70% 수준의 효율로 전기를 생산한다는 의미입니다. 단면 모듈에서는 계산에 반영되지 않습니다.\n\n"
        "추천값:\n"
        "- 일반 양면형 모듈: 65 ~ 75%\n"
        "- 고성능 양면형 모듈: 75 ~ 85%\n"
        "- 보수적인 추정: 60 ~ 70%\n"
        "- 현재 기본값: 70%"
    ),
    "azimuth": (
        "태양광 모듈이 바라보는 방향입니다. 이 코드에서는 0도=북쪽, 90도=동쪽, "
        "180도=남쪽, 270도=서쪽 기준으로 입력합니다.\n\n"
        "추천값:\n"
        "- 남향: 180도\n"
        "- 남동향: 135도\n"
        "- 남서향: 225도\n"
        "- 동향: 90도\n"
        "- 서향: 270도\n"
        "- 북향: 0도"
    ),
    "albedo": (
        "지표면이 햇빛을 얼마나 반사하는지를 나타내는 값입니다. 0에 가까울수록 반사가 적고, "
        "1에 가까울수록 반사가 많습니다. 양면형 모듈에서는 지면 반사광이 후면 발전량에 영향을 줍니다.\n\n"
        "추천값:\n"
        "- 아스팔트: 0.05 ~ 0.12\n"
        "- 흙/일반 지면: 0.15 ~ 0.25\n"
        "- 잔디: 0.20 ~ 0.30\n"
        "- 콘크리트: 0.25 ~ 0.40\n"
        "- 밝은 콘크리트/흰색 지붕: 0.40 ~ 0.60\n"
        "- 눈 덮인 지면: 0.60 ~ 0.90"
    ),
}

LOSS_HELP_TEXTS = {
    "soiling": (
        "모듈 표면에 먼지, 꽃가루, 새 배설물, 오염물 등이 쌓여 발생하는 손실입니다.\n\n"
        "추천값:\n"
        "- 관리가 잘 되는 설비: 1 ~ 2%\n"
        "- 일반적인 설비: 2 ~ 4%\n"
        "- 먼지가 많은 지역/청소가 드문 경우: 4 ~ 8%"
    ),
    "mismatch": (
        "모듈 간 출력 차이 때문에 발생하는 손실입니다. 같은 설비 안에서도 각 모듈의 성능이 조금씩 달라 "
        "전체 출력이 줄어들 수 있습니다.\n\n"
        "추천값:\n"
        "- 모듈 품질/배치가 좋은 경우: 0.5 ~ 1%\n"
        "- 일반적인 설비: 1 ~ 2%"
    ),
    "wiring": (
        "전선에서 전기가 이동하면서 발생하는 손실입니다. 전선 길이가 길거나 전선 굵기가 부족하면 손실이 커집니다.\n\n"
        "추천값:\n"
        "- 설계가 좋은 경우: 1% 내외\n"
        "- 일반적인 설비: 1 ~ 3%\n"
        "- 전선 거리가 긴 경우: 3% 이상"
    ),
    "connections": (
        "커넥터, 접속함, 단자 등 전기 연결부에서 발생하는 손실입니다.\n\n"
        "추천값:\n"
        "- 관리 상태가 좋은 경우: 0.5%\n"
        "- 일반적인 설비: 0.5 ~ 1%\n"
        "- 오래되었거나 접속부가 많은 경우: 1 ~ 2%"
    ),
    "lid": (
        "Light Induced Degradation의 약자로, 모듈이 처음 햇빛에 노출된 뒤 초기 성능이 약간 감소하는 현상입니다.\n\n"
        "추천값:\n"
        "- LID 저감 모듈: 0.5 ~ 1%\n"
        "- 일반 결정질 실리콘 모듈: 1 ~ 2%"
    ),
    "nameplate_rating": (
        "모듈에 표시된 정격출력과 실제 출력 사이의 차이로 인한 손실입니다.\n\n"
        "추천값:\n"
        "- 양의 출력공차 모듈 사용 시: 0%\n"
        "- 일반적인 설비: 0 ~ 1%\n"
        "- 보수적으로 계산할 때: 1%"
    ),
    "age": (
        "태양광 모듈이 시간이 지나면서 성능이 서서히 감소하는 손실입니다. 이 항목은 현재 계산에 바로 적용되는 "
        "누적 손실률로 입력하는 것이 자연스럽습니다.\n\n"
        "추천값:\n"
        "- 신규 설비: 0%\n"
        "- 설치 5년차: 약 2.5%\n"
        "- 설치 10년차: 약 5%\n"
        "- 일반적인 계산식: 사용연수 × 0.5%"
    ),
    "availability": (
        "고장, 점검, 인버터 정지, 계통 문제 등으로 설비가 정상 가동하지 못하는 시간에 따른 손실입니다.\n\n"
        "추천값:\n"
        "- 관리가 잘 되는 상업용 설비: 0.5 ~ 2%\n"
        "- 일반적인 설비: 1 ~ 3%\n"
        "- 정전/고장이 잦은 경우: 3% 이상"
    ),
}


def make_default_generator(name=None):
    return GeneratorConfig(
        name=name or "Generator 1",
        obstacles=[],
        losssettings=defaultlosssettings(),
    )


def clone_generator(gen):
    return GeneratorConfig(
        name=gen.name,
        modulepdcstcw=gen.modulepdcstcw,
        modulecount=gen.modulecount,
        invacratedwper=gen.invacratedwper,
        invertercount=gen.invertercount,
        etainvnom=gen.etainvnom,
        surfacetilt=gen.surfacetilt,
        surfaceazimuth=gen.surfaceazimuth,
        albedo=gen.albedo,
        mounting=gen.mounting,
        facetype=gen.facetype,
        bifacialityfactorpct=gen.bifacialityfactorpct,
        gammapctperc=gen.gammapctperc,
        plannedavailability=gen.plannedavailability,
        obstacles=copy.deepcopy(gen.obstacles or []),
        losssettings=copy.deepcopy(gen.losssettings or defaultlosssettings()),
    )


def load_form_from_generator(gen, edit_idx=None):
    st.session_state.form_version = st.session_state.get("form_version", 0) + 1
    purge_form_widget_state()
    st.session_state.show_loss_dialog = False
    st.session_state.form_edit_idx = edit_idx
    st.session_state.form_gen = clone_generator(gen)
    obstacles = [
        obs
        for obs in (gen.obstacles or [])
        if obs.get("enabled")
        or any(obs.get(k) not in (None, "", 0, 0.0) for k in ("centerazdeg", "distm", "heightm", "widthm"))
    ]
    st.session_state.form_obstacles = copy.deepcopy(obstacles)
    st.session_state.form_losssettings = copy.deepcopy(
        gen.losssettings or defaultlosssettings()
    )


def reset_form():
    st.session_state.form_version = st.session_state.get("form_version", 0) + 1
    purge_form_widget_state()
    st.session_state.show_loss_dialog = False
    st.session_state.form_edit_idx = None
    st.session_state.form_gen = make_default_generator(
        f"Generator {len(st.session_state.generators) + 1}"
    )
    st.session_state.form_obstacles = []
    st.session_state.form_losssettings = defaultlosssettings()


def refresh_form_widgets():
    st.session_state.form_version = st.session_state.get("form_version", 0) + 1


def add_form_obstacle():
    st.session_state.form_obstacles.append(
        {
            "enabled": True,
            "centerazdeg": 0.0,
            "distm": 0.0,
            "heightm": 0.0,
            "widthm": 0.0,
        }
    )
    refresh_form_widgets()


def delete_form_obstacle(idx):
    if 0 <= idx < len(st.session_state.form_obstacles):
        st.session_state.form_obstacles.pop(idx)
    refresh_form_widgets()


def is_old_default_obstacle(obs):
    default_obs = defaultobstacles()[0]
    return all(obs.get(key) == value for key, value in default_obs.items())


def purge_form_widget_state():
    for key in list(st.session_state.keys()):
        key_s = str(key)
        parts = key_s.split("_")
        is_versioned_form_key = (
            len(parts) > 2 and parts[0] == "form" and parts[1].isdigit()
        )
        if is_versioned_form_key:
            del st.session_state[key]


def first_existing_col(df, candidates):
    for col in candidates:
        if col in df.columns:
            return col
    return None

USAGE_DATE_COLUMN_CANDIDATES = [
    "Meter Reading Date",
    "검침일",
    "검침일시",
    "검침 날짜",
    "검침시간",
    "일시",
    "날짜시간",
]
USAGE_ENERGY_COLUMN_CANDIDATES = [
    "Forward Active Energy (kWh)",
    "순방향 유효전력량(KWH)",
    "순방향 유효전력량 (kWh)",
    "순방향유효전력량",
    "유효전력량",
    "전력사용량",
    "사용량",
    "kWh",
]
USAGE_METER_COLUMN_CANDIDATES = ["Meter Number", "계기번호", "계량기번호"]
USAGE_CUSTOMER_COLUMN_CANDIDATES = ["Customer Number", "고객번호", "수용가번호"]
WEEKDAY_ORDER = ["월", "화", "수", "목", "금", "토", "일"]
SEASON_GROUPS = {
    "하계": [7, 8],
    "춘추계": [3, 4, 5, 6, 9, 10],
    "동계": [11, 12, 1, 2],
}


def normalize_column_name(value):
    text = str(value).strip().lower()
    text = re.sub(r"[\s_\-./\\()\[\]{}]", "", text)
    text = text.replace("㎾h", "kwh")
    return text


def match_usage_column(columns, candidates):
    normalized_columns = {normalize_column_name(col): col for col in columns}
    for candidate in candidates:
        key = normalize_column_name(candidate)
        if key in normalized_columns:
            return normalized_columns[key]
    for candidate in candidates:
        key = normalize_column_name(candidate)
        for norm_col, original_col in normalized_columns.items():
            if key and (key in norm_col or norm_col in key):
                return original_col
    return None


def parse_usage_datetime(series):
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce")
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_datetime(series, unit="D", origin="1899-12-30", errors="coerce")
    raw = series.astype(str).str.strip()
    mask24 = raw.str.contains(r"24:00(?::00)?$", regex=True, na=False)
    normalized = raw.str.replace(r"24:00(?::00)?$", "00:00", regex=True)
    parsed = pd.to_datetime(normalized, errors="coerce")
    return parsed + pd.to_timedelta(mask24.astype(int), unit="D")


def parse_usage_energy(series):
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace(r"[^0-9.\-]", "", regex=True)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def read_uploaded_usage_file(uploaded_file):
    name = getattr(uploaded_file, "name", "uploaded_usage_file")
    suffix = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    raw = uploaded_file.getvalue()
    if suffix == "csv":
        last_error = None
        for encoding in ["utf-8-sig", "cp949", "euc-kr", "utf-8"]:
            try:
                return pd.read_csv(io.BytesIO(raw), encoding=encoding), name
            except Exception as exc:
                last_error = exc
        raise RuntimeError(f"CSV 파일을 읽지 못했습니다: {last_error}")
    if suffix in ("xlsx", "xls"):
        try:
            return pd.read_excel(io.BytesIO(raw), sheet_name=0), name
        except ImportError as exc:
            raise RuntimeError(
                "엑셀 파일을 읽는 데 필요한 라이브러리가 없습니다. "
                ".xlsx로 저장하거나 requirements.txt의 xlrd/openpyxl 설치를 확인해 주세요."
            ) from exc
    raise RuntimeError("지원하지 않는 파일 형식입니다. csv, xls, xlsx 파일을 업로드해 주세요.")


def build_usage_analysis(uploaded_file):
    raw_df, file_name = read_uploaded_usage_file(uploaded_file)
    df = raw_df.dropna(how="all").dropna(axis=1, how="all").copy()
    date_col = match_usage_column(df.columns, USAGE_DATE_COLUMN_CANDIDATES)
    energy_col = match_usage_column(df.columns, USAGE_ENERGY_COLUMN_CANDIDATES)
    meter_col = match_usage_column(df.columns, USAGE_METER_COLUMN_CANDIDATES)
    customer_col = match_usage_column(df.columns, USAGE_CUSTOMER_COLUMN_CANDIDATES)
    if date_col is None or energy_col is None:
        raise RuntimeError(
            "검침일 또는 전력량 컬럼을 자동 판독하지 못했습니다. "
            "예: Meter Reading Date/검침일, Forward Active Energy (kWh)/순방향 유효전력량(KWH)"
        )

    parsed_time = parse_usage_datetime(df[date_col])
    energy_kwh = parse_usage_energy(df[energy_col])
    parsed = pd.DataFrame(
        {
            "timestamp": parsed_time,
            "building_kwh": energy_kwh,
            "meter_number": df[meter_col].astype(str).str.strip() if meter_col else "",
            "customer_number": df[customer_col].astype(str).str.strip() if customer_col else "",
        }
    )
    bad_date_count = int(parsed["timestamp"].isna().sum())
    bad_energy_count = int(parsed["building_kwh"].isna().sum())
    parsed = parsed.dropna(subset=["timestamp", "building_kwh"]).sort_values("timestamp")
    parsed = parsed[parsed["building_kwh"] >= 0]
    if parsed.empty:
        raise RuntimeError("유효한 검침 데이터가 없습니다.")

    duplicate_count = int(parsed.duplicated(subset=["timestamp"]).sum())
    grouped = (
        parsed.groupby("timestamp", as_index=True)
        .agg(
            building_kwh=("building_kwh", "sum"),
            meter_number=("meter_number", "last"),
            customer_number=("customer_number", "last"),
        )
        .sort_index()
    )
    full_index = pd.date_range(grouped.index.min(), grouped.index.max(), freq="15min")
    usage_15min = grouped.reindex(full_index)
    usage_15min.index.name = "timestamp"
    usage_15min["has_usage_data"] = usage_15min["building_kwh"].notna()
    usage_15min["building_kw"] = usage_15min["building_kwh"] * 4.0
    missing_index = full_index.difference(grouped.index)

    meter_summary = pd.DataFrame()
    if meter_col:
        meter_summary = (
            parsed.groupby("meter_number")
            .agg(
                시작=("timestamp", "min"),
                종료=("timestamp", "max"),
                행수=("building_kwh", "size"),
                총사용량_kWh=("building_kwh", "sum"),
                최대15분_kWh=("building_kwh", "max"),
            )
            .reset_index()
            .rename(columns={"meter_number": "계기번호"})
        )
        meter_summary["최대수요_kW"] = meter_summary["최대15분_kWh"] * 4.0

    summary = {
        "파일명": file_name,
        "검침일 컬럼": date_col,
        "전력량 컬럼": energy_col,
        "계기번호 컬럼": meter_col or "없음",
        "고객번호 컬럼": customer_col or "없음",
        "시작": grouped.index.min(),
        "종료": grouped.index.max(),
        "원본 행 수": int(len(df)),
        "유효 행 수": int(len(parsed)),
        "예상 15분 슬롯": int(len(full_index)),
        "누락 15분 슬롯": int(len(missing_index)),
        "중복 시각 수": duplicate_count,
        "날짜 파싱 실패": bad_date_count,
        "전력량 파싱 실패": bad_energy_count,
        "총 사용량_kWh": float(grouped["building_kwh"].sum()),
        "최대 15분 사용량_kWh": float(grouped["building_kwh"].max()),
        "최대수요 환산_kW": float(grouped["building_kwh"].max() * 4.0),
    }
    return {
        "raw": df,
        "parsed": parsed,
        "usage_15min": usage_15min,
        "summary": summary,
        "meter_summary": meter_summary,
        "missing_index": missing_index,
    }


def usage_summary_frame(analysis):
    if not analysis:
        return pd.DataFrame()
    summary = analysis["summary"]
    rows = [
        ("파일명", summary["파일명"]),
        ("판독 검침일 컬럼", summary["검침일 컬럼"]),
        ("판독 전력량 컬럼", summary["전력량 컬럼"]),
        ("시작 시각", summary["시작"].strftime("%Y-%m-%d %H:%M")),
        ("종료 시각", summary["종료"].strftime("%Y-%m-%d %H:%M")),
        ("원본 행 수", f"{summary['원본 행 수']:,}"),
        ("유효 행 수", f"{summary['유효 행 수']:,}"),
        ("예상 15분 슬롯", f"{summary['예상 15분 슬롯']:,}"),
        ("누락 15분 슬롯", f"{summary['누락 15분 슬롯']:,}"),
        ("중복 시각 수", f"{summary['중복 시각 수']:,}"),
        ("총 사용량", f"{summary['총 사용량_kWh']:,.2f} kWh"),
        ("최대 15분 사용량", f"{summary['최대 15분 사용량_kWh']:,.2f} kWh"),
        ("최대수요 환산", f"{summary['최대수요 환산_kW']:,.2f} kW"),
    ]
    return pd.DataFrame(rows, columns=["항목", "값"])


def render_usage_data_summary():
    analysis = st.session_state.get("usage_analysis")
    st.subheader("전력사용량 데이터")
    if not analysis:
        st.info("사이드바에서 전력사용량 CSV/XLS/XLSX 파일을 업로드하면 요약이 표시됩니다.")
        return
    summary = analysis["summary"]
    if summary["누락 15분 슬롯"] > 0:
        st.warning("누락 시간대는 0으로 채우지 않고, 사용량 데이터가 있는 시각만 절감 분석에 반영합니다.")
    if summary["중복 시각 수"] > 0:
        st.info("중복 시각은 같은 시각의 전력량을 합산해 처리했습니다.")
    if summary["날짜 파싱 실패"] or summary["전력량 파싱 실패"]:
        st.warning(
            f"날짜 파싱 실패 {summary['날짜 파싱 실패']:,}건, "
            f"전력량 파싱 실패 {summary['전력량 파싱 실패']:,}건은 제외했습니다."
        )
    st.dataframe(usage_summary_frame(analysis), use_container_width=True, hide_index=True)
    meter_summary = analysis.get("meter_summary")
    if meter_summary is not None and not meter_summary.empty:
        st.markdown("##### 계기별 요약")
        st.dataframe(meter_summary, use_container_width=True, hide_index=True)


def resample_weather_to_usage_index(weatherhourly, target_index):
    if weatherhourly is None or weatherhourly.empty:
        return weatherhourly
    weather = weatherhourly.sort_index()
    target = pd.DatetimeIndex(target_index)
    if weather.index.tz is not None and target.tz is None:
        target = target.tz_localize(weather.index.tz)
    elif weather.index.tz is None and target.tz is not None:
        target = target.tz_localize(None)
    union_index = weather.index.union(target).sort_values()
    expanded = weather.reindex(union_index)
    numeric_cols = expanded.select_dtypes(include="number").columns
    if len(numeric_cols) > 0:
        expanded[numeric_cols] = expanded[numeric_cols].interpolate(method="time").ffill().bfill()
    other_cols = [col for col in expanded.columns if col not in numeric_cols]
    if other_cols:
        expanded[other_cols] = expanded[other_cols].ffill().bfill()
    return expanded.reindex(target)


def build_load_pv_match(usage_15min, pv_hourly):
    if usage_15min is None or pv_hourly is None or pv_hourly.empty:
        return None
    usage = usage_15min.copy()
    pv = pv_hourly[["generationkwh", "acpowerw"]].copy()
    if pv.index.tz is not None and usage.index.tz is None:
        usage.index = usage.index.tz_localize(pv.index.tz)
    elif pv.index.tz is None and usage.index.tz is not None:
        pv.index = pv.index.tz_localize(usage.index.tz)
    match = usage.join(pv.rename(columns={"generationkwh": "pv_kwh", "acpowerw": "pv_acpowerw"}), how="left")
    match["pv_kwh"] = pd.to_numeric(match["pv_kwh"], errors="coerce").fillna(0.0)
    match["pv_kw"] = pd.to_numeric(match["pv_acpowerw"], errors="coerce").fillna(0.0) / 1000.0
    match["building_kwh"] = pd.to_numeric(match["building_kwh"], errors="coerce")
    match["building_kw"] = match["building_kwh"] * 4.0
    match["has_usage_data"] = match["has_usage_data"].fillna(False).astype(bool)
    valid = match["has_usage_data"]
    match["self_consumed_kwh"] = float("nan")
    match["surplus_kwh"] = float("nan")
    match["net_load_kwh"] = float("nan")
    match.loc[valid, "self_consumed_kwh"] = match.loc[valid, ["building_kwh", "pv_kwh"]].min(axis=1)
    match.loc[valid, "surplus_kwh"] = (match.loc[valid, "pv_kwh"] - match.loc[valid, "building_kwh"]).clip(lower=0)
    match.loc[valid, "net_load_kwh"] = (match.loc[valid, "building_kwh"] - match.loc[valid, "pv_kwh"]).clip(lower=0)
    match["net_load_kw"] = match["net_load_kwh"] * 4.0
    return match


def compute_load_pv_kpis(match):
    if match is None or match.empty:
        return {}
    valid = match[match["has_usage_data"].fillna(False)].copy()
    if valid.empty:
        return {}
    total_usage = float(valid["building_kwh"].sum())
    pv_during_usage = float(valid["pv_kwh"].sum())
    total_self = float(valid["self_consumed_kwh"].sum())
    total_surplus = float(valid["surplus_kwh"].sum())
    peak_before_idx = valid["building_kw"].idxmax()
    peak_after_idx = valid["net_load_kw"].idxmax()
    peak_before = float(valid.loc[peak_before_idx, "building_kw"])
    peak_after = float(valid.loc[peak_after_idx, "net_load_kw"])
    return {
        "total_usage_kwh": total_usage,
        "pv_during_usage_kwh": pv_during_usage,
        "self_consumed_kwh": total_self,
        "surplus_kwh": total_surplus,
        "usage_reduction_pct": (total_self / total_usage * 100.0) if total_usage else 0.0,
        "self_consumption_pct": (total_self / pv_during_usage * 100.0) if pv_during_usage else 0.0,
        "surplus_pct": (total_surplus / pv_during_usage * 100.0) if pv_during_usage else 0.0,
        "peak_before_kw": peak_before,
        "peak_after_kw": peak_after,
        "peak_reduction_kw": peak_before - peak_after,
        "peak_reduction_pct": ((peak_before - peak_after) / peak_before * 100.0) if peak_before else 0.0,
        "peak_before_time": peak_before_idx,
        "peak_after_time": peak_after_idx,
    }


def season_name_for_month(month):
    for season, months in SEASON_GROUPS.items():
        if month in months:
            return season
    return "기타"


def hourly_pattern_table(match, value_col="building_kwh", season=None):
    if match is None or match.empty or value_col not in match.columns:
        return pd.DataFrame()
    data = match[match["has_usage_data"].fillna(False)][[value_col]].dropna().copy()
    if data.empty:
        return pd.DataFrame()
    if data.index.tz is not None:
        data.index = data.index.tz_convert(None)
    data["date"] = data.index.date
    data["hour"] = data.index.hour
    data["weekday"] = data.index.weekday
    data["season"] = [season_name_for_month(m) for m in data.index.month]
    if season:
        data = data[data["season"] == season]
    hourly = data.groupby(["date", "weekday", "hour"])[value_col].sum().reset_index()
    if hourly.empty:
        return pd.DataFrame()
    pattern = hourly.pivot_table(index="weekday", columns="hour", values=value_col, aggfunc="mean")
    pattern = pattern.reindex(index=range(7), columns=range(24))
    pattern.index = WEEKDAY_ORDER
    pattern.columns = [f"{hour:02d}시" for hour in pattern.columns]
    return pattern


def render_pattern_heatmap(pattern, title):
    if pattern is None or pattern.empty:
        st.info(f"{title} 데이터가 없습니다.")
        return
    plot_df = pattern.reset_index(names="요일").melt("요일", var_name="시간", value_name="평균 kWh")
    chart = (
        alt.Chart(plot_df)
        .mark_rect()
        .encode(
            x=alt.X("시간:N", sort=[f"{hour:02d}시" for hour in range(24)], title="시간"),
            y=alt.Y("요일:N", sort=WEEKDAY_ORDER, title="요일"),
            color=alt.Color("평균 kWh:Q", title="평균 kWh", scale=alt.Scale(scheme="yelloworangered")),
            tooltip=[
                alt.Tooltip("요일:N"),
                alt.Tooltip("시간:N"),
                alt.Tooltip("평균 kWh:Q", format=",.2f"),
            ],
        )
        .properties(height=260, title=title)
    )
    st.altair_chart(chart, use_container_width=True)
    st.dataframe(pattern.style.format("{:.2f}"), use_container_width=True)


def render_load_pv_analysis():
    match = st.session_state.get("load_pv_match")
    if match is None or match.empty:
        st.info("전력사용량 파일을 업로드한 뒤 계산하면 부하 절감 분석이 표시됩니다.")
        return
    kpis = compute_load_pv_kpis(match)
    if not kpis:
        st.info("사용 가능한 전력사용량 데이터가 없어 부하 절감 KPI를 계산하지 못했습니다.")
        return
    st.subheader("부하 절감 KPI")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("자료기간 총 사용량", f"{kpis['total_usage_kwh']:,.0f} kWh")
    c2.metric("자가소비 발전량", f"{kpis['self_consumed_kwh']:,.0f} kWh")
    c3.metric("전력량 절감률", f"{kpis['usage_reduction_pct']:,.1f}%")
    c4.metric("잉여 발전량", f"{kpis['surplus_kwh']:,.0f} kWh")
    c5, c6, c7, c8 = st.columns(4)
    c5.metric("기존 최대피크", f"{kpis['peak_before_kw']:,.1f} kW", kpis["peak_before_time"].strftime("%Y-%m-%d %H:%M"))
    c6.metric("PV 적용 후 최대피크", f"{kpis['peak_after_kw']:,.1f} kW", kpis["peak_after_time"].strftime("%Y-%m-%d %H:%M"))
    c7.metric("피크 절감", f"{kpis['peak_reduction_kw']:,.1f} kW")
    c8.metric("피크 절감률", f"{kpis['peak_reduction_pct']:,.1f}%")
    st.caption(f"자가소비율 {kpis['self_consumption_pct']:,.1f}%, 잉여율 {kpis['surplus_pct']:,.1f}%")

    chart_df = match[match["has_usage_data"].fillna(False)][["building_kw", "pv_kw", "net_load_kw"]].dropna().reset_index()
    chart_df.columns = ["time", "건물부하_kW", "태양광발전_kW", "순부하_kW"]
    long_df = chart_df.melt("time", var_name="구분", value_name="kW")
    chart = (
        alt.Chart(long_df)
        .mark_line()
        .encode(
            x=alt.X("time:T", title="시각"),
            y=alt.Y("kW:Q", title="kW"),
            color=alt.Color("구분:N", title="구분"),
            tooltip=[alt.Tooltip("time:T", title="시각"), alt.Tooltip("구분:N"), alt.Tooltip("kW:Q", format=",.1f")],
        )
        .properties(height=420)
        .interactive()
    )
    st.altair_chart(chart, use_container_width=True)

    monthly = match[match["has_usage_data"].fillna(False)][
        ["building_kwh", "pv_kwh", "self_consumed_kwh", "surplus_kwh", "net_load_kwh"]
    ].astype(float).resample("ME").sum()
    if not monthly.empty:
        st.markdown("##### 월별 요약")
        monthly_display = monthly.rename(
            columns={
                "building_kwh": "건물사용량_kWh",
                "pv_kwh": "태양광발전량_kWh",
                "self_consumed_kwh": "자가소비_kWh",
                "surplus_kwh": "잉여_kWh",
                "net_load_kwh": "순부하_kWh",
            }
        )
        st.dataframe(monthly_display.style.format("{:,.2f}"), use_container_width=True)

    st.subheader("요일·시간 부하패턴")
    pattern_target = st.radio("패턴 분석 대상", ["건물부하", "태양광 차감 후 순부하"], horizontal=True)
    value_col = "building_kwh" if pattern_target == "건물부하" else "net_load_kwh"
    render_pattern_heatmap(hourly_pattern_table(match, value_col=value_col), "연간 요일별 시간대 평균 사용패턴")
    season_tabs = st.tabs(list(SEASON_GROUPS.keys()))
    for season_tab, season in zip(season_tabs, SEASON_GROUPS.keys()):
        with season_tab:
            render_pattern_heatmap(hourly_pattern_table(match, value_col=value_col, season=season), f"{season} 요일별 시간대 평균 사용패턴")

def build_generator_from_form():
    base = st.session_state.form_gen
    return GeneratorConfig(
        name=base.name,
        modulepdcstcw=base.modulepdcstcw,
        modulecount=int(base.modulecount),
        invacratedwper=base.invacratedwper,
        invertercount=int(base.invertercount),
        etainvnom=base.etainvnom,
        surfacetilt=base.surfacetilt,
        surfaceazimuth=base.surfaceazimuth,
        albedo=base.albedo,
        mounting=base.mounting,
        facetype=base.facetype,
        bifacialityfactorpct=base.bifacialityfactorpct,
        gammapctperc=base.gammapctperc,
        plannedavailability=base.plannedavailability,
        obstacles=copy.deepcopy(st.session_state.form_obstacles),
        losssettings=copy.deepcopy(st.session_state.form_losssettings),
    )


def generator_summary_rows():
    rows = []
    for idx, gen in enumerate(st.session_state.generators, start=1):
        rows.append(
            {
                "번호": idx,
                "이름": gen.name,
                "모듈 STC [W]": gen.modulepdcstcw,
                "모듈 수": gen.modulecount,
                "DC 용량 [kW]": round(gen.pdc0totalw() / 1000.0, 2),
                "인버터 용량 [kW]": round(gen.invactotalw() / 1000.0, 2),
                "방위각 [deg]": gen.surfaceazimuth,
                "경사각 [deg]": gen.surfacetilt,
                "모듈 타입": gen.facetype,
                "장애물 수": sum(1 for obs in (gen.obstacles or []) if obs.get("enabled")),
            }
        )
    return rows


def build_generator_detail_summary(results=None):
    results = results or st.session_state.results or {}
    rows = []
    for gen in st.session_state.generators:
        generation_kwh = 0.0
        peak_kwh = 0.0
        if gen.name in results:
            hourly = results[gen.name]["hourly"]
            generation_kwh = float(hourly["generationkwh"].sum())
            peak_kwh = float(hourly["generationkwh"].max())
        rows.append(
            {
                "발전기": gen.name,
                "발전량 [kWh]": round(generation_kwh, 2),
                "DC 용량 [kW]": round(gen.pdc0totalw() / 1000.0, 2),
                "인버터 용량 [kW]": round(gen.invactotalw() / 1000.0, 2),
                "방위각 [deg]": round(float(gen.surfaceazimuth), 2),
                "경사각 [deg]": round(float(gen.surfacetilt), 2),
                "모듈 타입": gen.facetype,
            }
        )
    if rows:
        rows.append(
            {
                "발전기": "합계",
                "발전량 [kWh]": round(sum(row["발전량 [kWh]"] for row in rows), 2),
                "DC 용량 [kW]": round(sum(row["DC 용량 [kW]"] for row in rows), 2),
                "인버터 용량 [kW]": round(
                    sum(row["인버터 용량 [kW]"] for row in rows), 2
                ),
                "방위각 [deg]": "",
                "경사각 [deg]": "",
                "모듈 타입": "",
            }
        )
    return pd.DataFrame(rows)


def weather_export_frame(index):
    weatherhourly = st.session_state.get("weatherhourly")
    df = pd.DataFrame(index=index)
    if weatherhourly is None:
        return df
    columns = [
        ("shortwave_radiation", "전일사량_GHI_Wm2"),
        ("shortwaveradiation", "전일사량_GHI_Wm2"),
        ("direct_normal_irradiance", "직달일사량_DNI_Wm2"),
        ("directnormalirradiance", "직달일사량_DNI_Wm2"),
        ("diffuse_radiation", "산란일사량_DHI_Wm2"),
        ("diffuseradiation", "산란일사량_DHI_Wm2"),
        ("temperature_2m", "외기온도_C"),
        ("temperature2m", "외기온도_C"),
    ]
    used_labels = set()
    for source_col, label in columns:
        if source_col in weatherhourly.columns and label not in used_labels:
            df[label] = weatherhourly[source_col].reindex(index).values
            used_labels.add(label)
    return df


def clean_hourly_export_frame(name, hourly, include_weather=True):
    out = pd.DataFrame(index=hourly.index)
    out["발전기"] = name
    out["발전량_kWh"] = hourly["generationkwh"].values
    out["AC출력_kW"] = hourly["acpowerw"].values / 1000.0
    if "isshaded" in hourly.columns:
        out["음영여부"] = hourly["isshaded"].astype(bool).values
    if include_weather:
        out = pd.concat([out, weather_export_frame(hourly.index)], axis=1)
    try:
        out.index = out.index.tz_localize(None)
    except Exception:
        pass
    out.index.name = "날짜시각"
    return out


def clean_load_match_export(match):
    out = match.copy()
    try:
        out.index = out.index.tz_localize(None)
    except Exception:
        pass
    out.index.name = "날짜시각"
    out = out.reset_index()
    return out.rename(
        columns={
            "building_kwh": "건물사용량_kWh_15분",
            "building_kw": "건물부하_kW",
            "pv_kwh": "태양광발전량_kWh_15분",
            "pv_kw": "태양광발전_kW",
            "self_consumed_kwh": "자가소비_kWh",
            "surplus_kwh": "잉여발전량_kWh",
            "net_load_kwh": "순부하_kWh_15분",
            "net_load_kw": "순부하_kW",
            "has_usage_data": "사용량데이터_있음",
        }
    )


def build_csv_export(results):
    match = st.session_state.get("load_pv_match")
    if match is not None and not match.empty:
        csv_df = clean_load_match_export(match)
        return csv_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    frames = [
        clean_hourly_export_frame(name, data["hourly"], include_weather=True)
        for name, data in results.items()
    ]
    csv_df = pd.concat(frames).reset_index()
    return csv_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def style_worksheet(ws, freeze_cell="A2"):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[column_letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 28))
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
        ws.column_dimensions[column_letter].width = max_len + 2


def style_summary_worksheet(ws, detail_header_row=11):
    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, ws.max_column))
    for row_idx in range(3, 9):
        ws.cell(row_idx, 1).font = Font(bold=True)
        ws.cell(row_idx, 2).number_format = "#,##0.00"
    for cell in ws[detail_header_row]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{detail_header_row + 1}"
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[column_letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 28))
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
        ws.column_dimensions[column_letter].width = max_len + 2


def write_dataframe(ws, df, start_row=1, start_col=1, include_index=False):
    if include_index:
        df = df.reset_index()
    for col_offset, col_name in enumerate(df.columns):
        ws.cell(start_row, start_col + col_offset, col_name)
    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        for col_offset, value in enumerate(row):
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                if value.tzinfo is not None:
                    value = value.tz_convert(None)
                value = value.to_pydatetime()
            ws.cell(start_row + row_offset, start_col + col_offset, value)
    style_worksheet(ws)


def generate_excel_export(results, start_date, end_date, lat, lon):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"
    ws_summary["A1"] = "태양광 발전량 예측 결과"
    ws_summary["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("위도", lat),
        ("경도", lon),
        ("시작일", start_date.strftime("%Y-%m-%d")),
        ("종료일", end_date.strftime("%Y-%m-%d")),
        ("총 발전량 [kWh]", float(results["Total"]["hourly"]["generationkwh"].sum())),
        ("일평균 발전량 [kWh/day]", float(results["Total"]["daily"]["dailygenerationkwh"].mean())),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=3):
        ws_summary.cell(row_idx, 1, label)
        ws_summary.cell(row_idx, 2, value)
    detail_df = build_generator_detail_summary(results)
    for col_idx, col_name in enumerate(detail_df.columns, start=1):
        ws_summary.cell(11, col_idx, col_name)
    for row_offset, row in enumerate(detail_df.itertuples(index=False), start=12):
        for col_idx, value in enumerate(row, start=1):
            ws_summary.cell(row_offset, col_idx, value)
    style_summary_worksheet(ws_summary)

    total_hourly = clean_hourly_export_frame(
        "합계", results["Total"]["hourly"], include_weather=True
    ).drop(columns=["발전기"], errors="ignore")
    ws_total = wb.create_sheet("통합_시간별")
    write_dataframe(ws_total, total_hourly, include_index=True)

    generator_frames = []
    for name, data in results.items():
        if name == "Total":
            continue
        generator_frames.append(clean_hourly_export_frame(name, data["hourly"], include_weather=False))
    if generator_frames:
        generator_hourly = pd.concat(generator_frames).reset_index()
        ws_gen_hourly = wb.create_sheet("발전기별_시간별")
        write_dataframe(ws_gen_hourly, generator_hourly)

    daily_rows = []
    for name, data in results.items():
        daily = data["daily"].copy()
        try:
            daily.index = daily.index.tz_localize(None)
        except Exception:
            pass
        for ts, row in daily.iterrows():
            daily_rows.append(
                {
                    "날짜": ts.date() if hasattr(ts, "date") else ts,
                    "발전기": "합계" if name == "Total" else name,
                    "일 발전량_kWh": float(row["dailygenerationkwh"]),
                }
            )
    ws_daily = wb.create_sheet("일별_요약")
    write_dataframe(ws_daily, pd.DataFrame(daily_rows))

    match = st.session_state.get("load_pv_match")
    if match is not None and not match.empty:
        kpis = compute_load_pv_kpis(match)
        if kpis:
            start_row = ws_summary.max_row + 3
            ws_summary.cell(start_row, 1, "부하 절감 분석")
            ws_summary.cell(start_row, 1).font = Font(bold=True)
            load_rows = [
                ("자료기간 총 사용량 [kWh]", kpis["total_usage_kwh"]),
                ("태양광 자가소비 [kWh]", kpis["self_consumed_kwh"]),
                ("잉여 발전량 [kWh]", kpis["surplus_kwh"]),
                ("전력량 절감률 [%]", kpis["usage_reduction_pct"]),
                ("기존 최대피크 [kW]", kpis["peak_before_kw"]),
                ("PV 적용 후 최대피크 [kW]", kpis["peak_after_kw"]),
                ("피크 절감 [kW]", kpis["peak_reduction_kw"]),
                ("피크 절감률 [%]", kpis["peak_reduction_pct"]),
            ]
            for row_idx, (label, value) in enumerate(load_rows, start=start_row + 1):
                ws_summary.cell(row_idx, 1, label)
                ws_summary.cell(row_idx, 2, value)

        match_export = clean_load_match_export(match)
        ws_match = wb.create_sheet("15분_부하_PV_매칭")
        write_dataframe(ws_match, match_export)

        monthly = match[match["has_usage_data"].fillna(False)][
            ["building_kwh", "pv_kwh", "self_consumed_kwh", "surplus_kwh", "net_load_kwh"]
        ].astype(float).resample("ME").sum().rename(
            columns={
                "building_kwh": "건물사용량_kWh",
                "pv_kwh": "태양광발전량_kWh",
                "self_consumed_kwh": "자가소비_kWh",
                "surplus_kwh": "잉여_kWh",
                "net_load_kwh": "순부하_kWh",
            }
        )
        ws_monthly = wb.create_sheet("월별_요약")
        write_dataframe(ws_monthly, monthly, include_index=True)

        pattern_items = [("연간_요일시간패턴", None)] + [(f"{season}_요일시간패턴", season) for season in SEASON_GROUPS]
        for sheet_name, season in pattern_items:
            pattern = hourly_pattern_table(match, value_col="building_kwh", season=season)
            if not pattern.empty:
                ws_pattern = wb.create_sheet(sheet_name[:31])
                write_dataframe(ws_pattern, pattern, include_index=True)

    usage_analysis = st.session_state.get("usage_analysis")
    if usage_analysis:
        missing_index = usage_analysis.get("missing_index")
        if missing_index is not None and len(missing_index) > 0:
            missing_df = pd.DataFrame({"누락시각": missing_index.astype(str)})
            ws_missing = wb.create_sheet("누락데이터_요약")
            write_dataframe(ws_missing, missing_df)

    settings_df = pd.DataFrame(generator_summary_rows())
    ws_settings = wb.create_sheet("발전기_설정")
    write_dataframe(ws_settings, settings_df)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def render_hourly_result_chart(hourly, weatherhourly, ghi_col, temp_col):
    chart_df = pd.DataFrame(
        {
            "time": hourly.index,
            "generationkwh": hourly["generationkwh"].values,
        }
    )
    chart = (
        alt.Chart(chart_df)
        .mark_area(opacity=0.35, color="#1F77B4")
        .encode(
            x=alt.X("time:T", title="시각"),
            y=alt.Y("generationkwh:Q", title="발전량 [kWh]", scale=alt.Scale(zero=True)),
            tooltip=[
                alt.Tooltip("time:T", title="시각"),
                alt.Tooltip("generationkwh:Q", title="발전량 [kWh]", format=",.2f"),
            ],
        )
        .properties(height=360)
        .interactive()
    )
    st.altair_chart(chart, use_container_width=True)

@st.dialog("손실 설정")
def render_loss_dialog(key_prefix):
    loss_labels = [
        ("soiling", "Soiling"),
        ("mismatch", "Mismatch"),
        ("wiring", "Wiring"),
        ("connections", "Connections"),
        ("lid", "LID"),
        ("nameplate_rating", "Nameplate"),
        ("age", "Age"),
        ("availability", "Availability loss"),
    ]
    for row_start in range(0, len(loss_labels), 2):
        row_cols = st.columns(2)
        for col, (key, label) in zip(row_cols, loss_labels[row_start:row_start + 2]):
            cur = st.session_state.form_losssettings.get(
                key, {"enabled": True, "value": float(LOSSPARAMS.get(key, 0.0))}
            )
            with col:
                enabled = st.checkbox(
                    label,
                    value=bool(cur.get("enabled", True)),
                    key=f"{key_prefix}_loss_dialog_en_{key}",
                    help=LOSS_HELP_TEXTS.get(key),
                )
                value = st.number_input(
                    "%",
                    value=float(cur.get("value", 0.0)),
                    disabled=not enabled,
                    key=f"{key_prefix}_loss_dialog_val_{key}",
                    label_visibility="collapsed",
                )
            st.session_state.form_losssettings[key] = {
                "enabled": enabled,
                "value": value,
            }

    if st.button("완료", type="primary", use_container_width=True):
        st.session_state.show_loss_dialog = False
        st.rerun()


def render_generator_input():
    gen = st.session_state.form_gen
    key_prefix = f"form_{st.session_state.get('form_version', 0)}"

    st.subheader("발전기 정보 입력")
    mode_text = (
        f"{st.session_state.form_edit_idx + 1}번 발전기 수정 중"
        if st.session_state.form_edit_idx is not None
        else "새 발전기 등록"
    )
    st.caption(mode_text)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("##### 기본 정보")
        gen.name = st.text_input("이름", value=gen.name, key=f"{key_prefix}_name")

        st.markdown("##### 모듈 설정")
        gen.modulepdcstcw = st.number_input(
            "모듈 STC [W]",
            value=float(gen.modulepdcstcw),
            min_value=0.0,
            step=10.0,
            key=f"{key_prefix}_modulepdcstcw",
        )
        gen.modulecount = st.number_input(
            "모듈 개수",
            value=int(gen.modulecount),
            min_value=1,
            step=1,
            key=f"{key_prefix}_modulecount",
        )
        gen.gammapctperc = st.number_input(
            "온도계수 Gamma [%/°C]",
            value=float(gen.gammapctperc),
            step=0.01,
            key=f"{key_prefix}_gamma",
            help=FIELD_HELP_TEXTS["gamma"],
        )
        face_options = ["Monofacial", "Bifacial"]
        face_idx = face_options.index(gen.facetype) if gen.facetype in face_options else 1
        gen.facetype = st.selectbox(
            "모듈 타입", face_options, index=face_idx, key=f"{key_prefix}_facetype"
        )
        if gen.facetype == "Bifacial":
            gen.bifacialityfactorpct = st.number_input(
                "양면 발전 계수 [%]",
                value=float(getattr(gen, "bifacialityfactorpct", 70.0)),
                min_value=0.0,
                max_value=100.0,
                step=1.0,
                key=f"{key_prefix}_bifaciality",
                help=FIELD_HELP_TEXTS["bifaciality"],
            )

    with col2:
        st.markdown("##### 인버터 설정")
        gen.invacratedwper = (
            st.number_input(
                "인버터 1대 용량 [kW]",
                value=float(gen.invacratedwper) / 1000.0,
                min_value=0.0,
                step=0.1,
                key=f"{key_prefix}_invkw",
            )
            * 1000.0
        )
        gen.invertercount = st.number_input(
            "인버터 개수",
            value=int(gen.invertercount),
            min_value=1,
            step=1,
            key=f"{key_prefix}_invertercount",
        )
        gen.etainvnom = (
            st.number_input(
                "인버터 효율 [%]",
                value=float(gen.etainvnom) * 100.0,
                min_value=0.0,
                max_value=100.0,
                step=0.1,
                key=f"{key_prefix}_eta",
            )
            / 100.0
        )

        st.markdown("##### 설치 환경")
        gen.surfaceazimuth = st.number_input(
            "방위각 [deg]",
            value=float(gen.surfaceazimuth),
            step=1.0,
            key=f"{key_prefix}_surfaceazimuth",
            help=FIELD_HELP_TEXTS["azimuth"],
        )
        gen.surfacetilt = st.number_input(
            "경사각 [deg]",
            value=float(gen.surfacetilt),
            min_value=0.0,
            step=0.1,
            key=f"{key_prefix}_surfacetilt",
        )
        mount_options = ["Open rack", "Close mount"]
        mount_idx = mount_options.index(gen.mounting) if gen.mounting in mount_options else 0
        gen.mounting = st.selectbox(
            "설치 형태", mount_options, index=mount_idx, key=f"{key_prefix}_mounting"
        )
        gen.albedo = st.number_input(
            "알베도",
            value=float(gen.albedo),
            min_value=0.0,
            max_value=1.0,
            step=0.01,
            key=f"{key_prefix}_albedo",
            help=FIELD_HELP_TEXTS["albedo"],
        )
        gen.plannedavailability = (
            st.number_input(
                "계획 가동률 [%]",
                value=float(gen.plannedavailability) * 100.0,
                min_value=0.0,
                max_value=100.0,
                step=0.1,
                key=f"{key_prefix}_plannedavailability",
            )
            / 100.0
        )

    with col3:
        st.markdown("##### 장애물 및 손실 설정")
        with st.expander("장애물", expanded=True):
            if not st.session_state.form_obstacles:
                st.caption("등록된 장애물이 없습니다.")

            if st.button(
                "+ 장애물 추가",
                key=f"{key_prefix}_add_obstacle",
                use_container_width=True,
            ):
                st.session_state.form_obstacles.append(
                    {
                        "enabled": True,
                        "centerazdeg": 0.0,
                        "distm": 0.0,
                        "heightm": 0.0,
                        "widthm": 0.0,
                    }
                )
                st.rerun()

            for idx, obs in enumerate(st.session_state.form_obstacles):
                st.divider()
                st.markdown(f"**장애물 {idx + 1}**")
                obs["enabled"] = True
                c1, c2 = st.columns(2)
                obs["centerazdeg"] = c1.number_input(
                    "방위각 [deg]",
                    value=float(obs.get("centerazdeg") or 0.0),
                    key=f"{key_prefix}_obs_az_{idx}",
                )
                obs["distm"] = c2.number_input(
                    "거리 [m]",
                    value=float(obs.get("distm") or 0.0),
                    key=f"{key_prefix}_obs_dist_{idx}",
                )
                c3, c4 = st.columns(2)
                obs["heightm"] = c3.number_input(
                    "높이 [m]",
                    value=float(obs.get("heightm") or 0.0),
                    key=f"{key_prefix}_obs_height_{idx}",
                )
                obs["widthm"] = c4.number_input(
                    "너비 [m]",
                    value=float(obs.get("widthm") or 0.0),
                    key=f"{key_prefix}_obs_width_{idx}",
                )
                if st.button("장애물 삭제", key=f"{key_prefix}_delete_obstacle_{idx}"):
                    st.session_state.form_obstacles.pop(idx)
                    st.rerun()

        if st.button("손실 설정", key=f"{key_prefix}_open_loss_dialog", use_container_width=True):
            st.session_state.show_loss_dialog = True
            st.rerun()

        if st.session_state.get("show_loss_dialog", False):
            render_loss_dialog(key_prefix)

    st.divider()
    c_submit, c_reset = st.columns([2, 1])
    submit_label = "선택 항목 수정 완료" if st.session_state.form_edit_idx is not None else "+ 목록에 추가"
    if c_submit.button(submit_label, type="primary", use_container_width=True):
        new_gen = build_generator_from_form()
        if not new_gen.name.strip():
            st.error("발전기 이름을 입력해 주세요.")
        elif st.session_state.form_edit_idx is None:
            st.session_state.generators.append(new_gen)
            st.success(f"{new_gen.name} 발전기를 계산 대상 목록에 추가했습니다.")
            reset_form()
            st.rerun()
        else:
            st.session_state.generators[st.session_state.form_edit_idx] = new_gen
            st.success(f"{new_gen.name} 발전기 정보를 수정했습니다.")
            reset_form()
            st.rerun()

    if c_reset.button("입력 초기화", use_container_width=True):
        reset_form()
        st.rerun()


def render_results(start_date, end_date, lat, lon):
    if st.session_state.results is None:
        st.info("계산 대상 목록에서 시뮬레이션을 실행하면 결과가 여기에 표시됩니다.")
        return

    if st.session_state.get("capped_message"):
        st.warning(
            "Open-Meteo 예보 데이터 제공 범위 때문에 "
            f"{st.session_state.calc_end_date.strftime('%Y-%m-%d')}까지 계산했습니다."
        )

    st.subheader("시뮬레이션 결과")
    result_keys = list(st.session_state.results.keys())
    if "Total" in result_keys:
        result_keys = ["Total"] + [key for key in result_keys if key != "Total"]
    tabs = st.tabs(result_keys)

    weatherhourly = st.session_state.get("weatherhourly")
    ghi_col = first_existing_col(
        weatherhourly, ["shortwave_radiation", "shortwaveradiation"]
    ) if weatherhourly is not None else None
    temp_col = first_existing_col(
        weatherhourly, ["temperature_2m", "temperature2m"]
    ) if weatherhourly is not None else None

    for tab, key in zip(tabs, result_keys):
        with tab:
            hourly = st.session_state.results[key]["hourly"]
            daily = st.session_state.results[key]["daily"]
            total_kwh = daily["dailygenerationkwh"].sum()
            hourly_generation = hourly["generationkwh"]
            peak_kwh = float(hourly_generation.max())
            peak_time = hourly_generation.idxmax()
            avg_daily_kwh = float(daily["dailygenerationkwh"].mean())

            kpi1, kpi2, kpi3 = st.columns(3)
            kpi1.metric("총 발전량", f"{total_kwh:,.2f} kWh")
            kpi2.metric(
                "최대 발전량",
                f"{peak_kwh:,.2f} kWh",
                peak_time.strftime("%Y-%m-%d %H:%M"),
            )
            kpi3.metric("일평균 발전량", f"{avg_daily_kwh:,.2f} kWh/day")

            render_hourly_result_chart(hourly, weatherhourly, ghi_col, temp_col)

            st.markdown("##### 날짜별 시간대 발전량 [kWh]")
            df = hourly["generationkwh"].copy()
            try:
                df.index = df.index.tz_localize(None)
            except Exception:
                pass
            table_df = df.reset_index()
            table_df.columns = ["time", "generationkwh"]
            table_df["날짜"] = table_df["time"].dt.strftime("%Y-%m-%d")
            table_df["시간"] = table_df["time"].dt.hour
            mat = table_df.pivot_table(
                index="날짜",
                columns="시간",
                values="generationkwh",
                aggfunc="sum",
            ).reindex(columns=range(24)).fillna(0.0)
            mat.columns = [f"{hour:02d}시" for hour in mat.columns]
            st.dataframe(
                mat.style.format("{:.2f}").background_gradient(cmap="YlOrRd", axis=None),
                use_container_width=True,
                height=min(420, 74 + 35 * len(mat)),
            )

            st.markdown("##### 발전기별 세부 요약")
            detail_df = build_generator_detail_summary(st.session_state.results)
            st.dataframe(
                detail_df,
                use_container_width=True,
                hide_index=True,
            )

    st.divider()
    render_load_pv_analysis()

    st.divider()
    st.subheader("데이터 다운로드")
    excel_output = generate_excel_export(
        st.session_state.results, start_date, end_date, lat, lon
    )
    csv_output = build_csv_export(st.session_state.results)
    col_excel, col_csv = st.columns(2)
    with col_excel:
        st.download_button(
            label="Excel 결과 파일 다운로드",
            data=excel_output.getvalue(),
            file_name=f"PV_Forecast_{start_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with col_csv:
        st.download_button(
            label="CSV 시간별 데이터 다운로드",
            data=csv_output,
            file_name=f"PV_Forecast_Hourly_{start_date.strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
        )


if "generators" not in st.session_state:
    st.session_state.generators = []
if "results" not in st.session_state:
    st.session_state.results = None
if "weatherhourly" not in st.session_state:
    st.session_state.weatherhourly = None
if "capped_message" not in st.session_state:
    st.session_state.capped_message = False
if "calc_end_date" not in st.session_state:
    st.session_state.calc_end_date = None
if "usage_analysis" not in st.session_state:
    st.session_state.usage_analysis = None
if "usage_error" not in st.session_state:
    st.session_state.usage_error = None
if "usage_file_id" not in st.session_state:
    st.session_state.usage_file_id = None
if "load_pv_match" not in st.session_state:
    st.session_state.load_pv_match = None
if "form_gen" not in st.session_state:
    reset_form()
if not st.session_state.get("empty_obstacle_default_applied", False):
    form_obstacles = st.session_state.get("form_obstacles", [])
    if len(form_obstacles) == 1 and is_old_default_obstacle(form_obstacles[0]):
        st.session_state.form_obstacles = []
        refresh_form_widgets()
    st.session_state.empty_obstacle_default_applied = True


with st.sidebar:
    st.header("위치 설정")
    lat = st.number_input("위도", value=37.4317862, format="%.7f")
    lon = st.number_input("경도", value=126.6485109, format="%.7f")

    st.divider()
    st.header("전력사용량")
    usage_file = st.file_uploader(
        "전력사용량 파일",
        type=["csv", "xls", "xlsx"],
        help="검침일과 전력량 컬럼을 자동 판독합니다.",
    )
    if usage_file is None:
        st.session_state.usage_analysis = None
        st.session_state.usage_error = None
        st.session_state.usage_file_id = None
        st.session_state.load_pv_match = None
        st.caption("업로드된 전력사용량 파일이 없습니다.")
    else:
        file_id = (usage_file.name, getattr(usage_file, "size", None))
        if st.session_state.usage_file_id != file_id:
            try:
                st.session_state.usage_analysis = build_usage_analysis(usage_file)
                st.session_state.usage_error = None
                st.session_state.usage_file_id = file_id
                st.session_state.load_pv_match = None
            except Exception as exc:
                st.session_state.usage_analysis = None
                st.session_state.usage_error = str(exc)
                st.session_state.usage_file_id = file_id
                st.session_state.load_pv_match = None
        if st.session_state.usage_error:
            st.error(st.session_state.usage_error)
        elif st.session_state.usage_analysis:
            usage_summary = st.session_state.usage_analysis["summary"]
            st.success("전력사용량 파일 업로드됨")
            st.caption(
                f"{usage_summary['시작'].strftime('%Y-%m-%d')} ~ "
                f"{usage_summary['종료'].strftime('%Y-%m-%d')} · 15분 데이터"
            )

    st.divider()
    st.header("기간 설정")
    today = datetime.date.today()
    usage_analysis = st.session_state.get("usage_analysis")
    default_start = usage_analysis["summary"]["시작"].date() if usage_analysis else today
    default_end = usage_analysis["summary"]["종료"].date() if usage_analysis else today + datetime.timedelta(days=7)
    start_date = st.date_input("시작일", default_start)
    end_date = st.date_input("종료일", default_end)
    st.caption("전력사용량 파일이 있으면 해당 기간을 기본값으로 사용합니다.")


st.title("PV-Forecast")
st.caption("태양광 발전량 예측 및 발전기별 시뮬레이션")

tab_input, tab_list, tab_result = st.tabs(
    ["발전기 정보 입력", "계산 대상 목록", "시뮬레이션 결과"]
)

with tab_input:
    render_generator_input()

with tab_list:
    st.subheader("계산 대상 목록")
    render_usage_data_summary()
    st.divider()
    if not st.session_state.generators:
        st.info("아직 등록된 발전기가 없습니다. '발전기 정보 입력' 탭에서 발전기를 추가해 주세요.")
    else:
        summary = pd.DataFrame(generator_summary_rows())
        st.dataframe(summary, use_container_width=True, hide_index=True)

        st.divider()
        c_select, c_edit, c_delete = st.columns([1.2, 1, 1])
        with c_select:
            label_col, input_col = st.columns([0.9, 1.1], vertical_alignment="center")
            label_col.markdown("선택 번호")
            selected_no = input_col.number_input(
                "선택 번호",
                min_value=1,
                max_value=len(st.session_state.generators),
                value=1,
                step=1,
                label_visibility="collapsed",
            )
        selected_idx = int(selected_no) - 1

        if c_edit.button("선택 수정", use_container_width=True):
            load_form_from_generator(
                st.session_state.generators[selected_idx],
                edit_idx=selected_idx,
            )
            st.success(
                f"{selected_no}번 발전기를 수정 모드로 전환했습니다. "
                "'발전기 정보 입력' 탭을 확인해 주세요."
            )
            st.rerun()

        if c_delete.button("선택 삭제", type="secondary", use_container_width=True):
            deleted_name = st.session_state.generators[selected_idx].name
            st.session_state.generators.pop(selected_idx)
            if st.session_state.form_edit_idx == selected_idx:
                reset_form()
            st.success(f"{deleted_name} 발전기를 삭제했습니다.")
            st.rerun()

        st.divider()
        if st.button("전체 시뮬레이션 계산 실행", type="primary", use_container_width=True):
            if start_date > end_date:
                st.error("종료일은 시작일 이후여야 합니다.")
            else:
                max_forecast_date = today + datetime.timedelta(days=15)
                capped = False
                calc_end_date = end_date
                if end_date > max_forecast_date:
                    calc_end_date = max_forecast_date
                    capped = True

                st.session_state.capped_message = capped
                st.session_state.calc_end_date = calc_end_date

                with st.status("태양광 발전량 시뮬레이션 진행 중...", expanded=True) as status:
                    try:
                        start_s = start_date.strftime("%Y-%m-%d")
                        end_s = calc_end_date.strftime("%Y-%m-%d")

                        status.write("Open-Meteo 기상 데이터를 수집하는 중...")
                        tz, altitude_m = resolvetimezoneandelevation(lat, lon)
                        weatherhourly, weatherdaily = getweatherdatamixed(
                            lat, lon, start_s, end_s, tz
                        )

                        usage_for_calc = None
                        weather_for_pv = weatherhourly
                        if st.session_state.get("usage_analysis"):
                            usage_full = st.session_state.usage_analysis["usage_15min"]
                            period_start = pd.Timestamp(start_date)
                            period_end = pd.Timestamp(calc_end_date) + pd.Timedelta(days=1) - pd.Timedelta(minutes=15)
                            usage_for_calc = usage_full.loc[
                                (usage_full.index >= period_start) & (usage_full.index <= period_end)
                            ]
                            if not usage_for_calc.empty:
                                status.write("전력사용량 15분 시각에 맞춰 기상 데이터를 변환하는 중...")
                                weather_for_pv = resample_weather_to_usage_index(weatherhourly, usage_for_calc.index)

                        results = {}
                        total_hourly = None

                        for idx, gen in enumerate(st.session_state.generators, start=1):
                            status.write(f"{idx}. [{gen.name}] 발전량 계산 중...")
                            obstaclesenabled = [
                                obs for obs in gen.obstacles if obs.get("enabled")
                            ]
                            lossparams = lossparamsforgenerator(gen)
                            hourly, daily = computepvforgenerator(
                                weather_for_pv,
                                gen,
                                obstaclesenabled,
                                lossparams,
                                lat,
                                lon,
                                tz,
                                altitude_m,
                            )
                            results[gen.name] = {"hourly": hourly, "daily": daily}

                            cur = hourly[["acpowerw", "generationkwh"]].copy()
                            total_hourly = cur if total_hourly is None else total_hourly.add(cur, fill_value=0.0)

                        status.write("전체 발전기 합계를 계산하는 중...")
                        total_daily = total_hourly["generationkwh"].resample("D").sum().to_frame(
                            name="dailygenerationkwh"
                        )
                        results["Total"] = {"hourly": total_hourly, "daily": total_daily}

                        st.session_state.results = results
                        st.session_state.weatherhourly = weather_for_pv
                        st.session_state.load_pv_match = build_load_pv_match(usage_for_calc, total_hourly) if usage_for_calc is not None else None
                        status.update(label="계산 완료", state="complete", expanded=False)
                        st.success("'시뮬레이션 결과' 탭에서 결과를 확인해 주세요.")
                    except Exception as exc:
                        status.update(label="오류 발생", state="error", expanded=True)
                        st.error(f"오류 발생: {exc}")

with tab_result:
    render_results(start_date, end_date, lat, lon)
