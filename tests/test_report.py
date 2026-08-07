"""Excel 출력과 CLI 배치 (요구사항서 10.3·10.4).

산출물은 숫자만으로 읽히지 않는다. **무엇을 근거로 계산했고(5.8), 무엇을 빼놓았고
(5.1), 무엇을 조심해야 하며(9.4), 어디까지 믿을 수 있는지(부록 D)** 가 같은 장에
있어야 한다. 이 파일은 그 문구들이 실제로 실리는지를 못 박는다.

tz 해제와 파일명 접미사는 Windows·Excel 환경에서 매번 걸려 넘어지던 곳이다.
pvlib 결과는 항상 tz-aware 이고, Excel 이 파일을 열고 있으면 덮어쓰기가 막힌다.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from kwise.cli import app
from kwise.compare import CombinationSpec, ComparisonResult, evaluate_combination
from kwise.diagnose import Diagnosis
from kwise.io import UsageData
from kwise.measures import Certainty, EssResult, TariffSwitchResult, evaluate_contract_adjustment
from kwise.quality import QualityReport
from kwise.report import (
    CONTRACT_CHANGE_WARNING,
    KNOWN_LIMITS,
    NO_PV_SENSITIVITY_NOTE,
    NOT_INCLUDED_NOTICE,
    SHEET_ORDER,
    UNPRICED_REASONS,
    ReportSections,
    ReportWriteError,
    build_sheets,
    export_report,
    format_won,
    load_batch_config,
    measure_summary_frame,
    no_pv_sensitivity_frame,
    result_path,
    run_batch,
    strip_timezone,
    write_workbook,
)
from kwise.tariff import BillingResult, TariffSelection, TariffTable
from tests._synthetic import write_month

CURRENT = TariffSelection("general_b", "high_a", "I")
BEST = TariffSelection("general_b", "high_a", "II")


# --------------------------------------------------------------------- 픽스처


@pytest.fixture(scope="session")
def sample_measure_rows(sample_switch: TariffSwitchResult, sample_ess: EssResult) -> pd.DataFrame:
    """수단별 결과 시트. 계약전력 조정은 하한 미확인이라 금액이 비는 경로다."""
    return measure_summary_frame(switch=sample_switch, ess=sample_ess)


@pytest.fixture(scope="session")
def sample_sections(
    sample_usage: UsageData,
    sample_bill: BillingResult,
    sample_diagnosis: Diagnosis,
    sample_comparison: ComparisonResult,
    sample_measure_rows: pd.DataFrame,
) -> ReportSections:
    """여덟 장이 모두 채워지는 케이스."""
    return ReportSections(
        usage=sample_usage,
        bill=sample_bill,
        diagnosis=sample_diagnosis,
        comparison=sample_comparison,
        sensitivity=no_pv_sensitivity_frame(),
        measure_rows=sample_measure_rows,
    )


@pytest.fixture(scope="session")
def sample_sheets(sample_sections: ReportSections) -> dict[str, pd.DataFrame]:
    return build_sheets(sample_sections)


@pytest.fixture(scope="session")
def summary_text(sample_sheets: dict[str, pd.DataFrame]) -> str:
    """요약 시트를 한 덩어리 문자열로. 문구 존재 확인에 쓴다."""
    summary = sample_sheets["요약"].reset_index()
    return "\n".join(
        " | ".join(str(value) for value in row) for row in summary.itertuples(index=False)
    )


# --------------------------------------------------------------------- 시트 구성


def test_workbook_has_the_eight_sheets_in_order(sample_sheets: dict[str, pd.DataFrame]) -> None:
    """요약 / 진단 / 월별 집계 / 15분 시계열 / 요금 계산 명세 / 수단별 결과 / 조합 비교 / 감도."""
    assert tuple(sample_sheets) == SHEET_ORDER
    assert len(SHEET_ORDER) == 8


def test_sheet_order_survives_the_round_trip(
    sample_sections: ReportSections, tmp_path: Path
) -> None:
    """openpyxl 로 다시 열어도 여덟 장이 같은 순서다."""
    path = export_report(sample_sections, output_dir=tmp_path)
    assert path.is_file()
    workbook = load_workbook(path, read_only=True)
    try:
        assert tuple(workbook.sheetnames) == SHEET_ORDER
    finally:
        workbook.close()


def test_timeseries_sheet_carries_every_slot(
    sample_sheets: dict[str, pd.DataFrame], sample_usage: UsageData
) -> None:
    """35,328 행이 통째로 실린다. 결측은 미보간이므로 표시만 한다."""
    timeseries = sample_sheets["15분 시계열"]
    assert len(timeseries) == len(sample_usage.kw)
    assert list(timeseries.columns) == ["kw", "kwh", "결측"]
    assert int(timeseries["결측"].sum()) == sample_usage.meta.missing_rows


def test_optional_sheets_are_absent_not_blank(
    sample_usage: UsageData, sample_bill: BillingResult
) -> None:
    """진단·수단·조합·감도가 없으면 시트를 만들지 않는다. 빈 장을 남기지 않는다."""
    sheets = build_sheets(ReportSections(usage=sample_usage, bill=sample_bill))
    assert tuple(sheets) == ("요약", "월별 집계", "15분 시계열", "요금 계산 명세")


def test_timeseries_can_be_switched_off(
    sample_usage: UsageData, sample_bill: BillingResult
) -> None:
    """케이스가 많을 때 35,328 행 × N 을 피할 수 있어야 한다."""
    sheets = build_sheets(
        ReportSections(usage=sample_usage, bill=sample_bill, include_timeseries=False)
    )
    assert "15분 시계열" not in sheets


def test_monthly_sheet_shows_both_demand_columns(
    sample_sheets: dict[str, pd.DataFrame],
) -> None:
    """관측 최대와 요금적용전력을 함께 싣는다. 둘은 다른 값이다 (5.2)."""
    monthly = sample_sheets["월별 집계"]
    assert {"max_demand_kw", "billing_demand_kw", "missing_ratio"} <= set(monthly.columns)
    august = monthly.loc["2023-08"]
    assert august["max_demand_kw"] < august["billing_demand_kw"]


# --------------------------------------------------------------------- tz 해제


def _tz_frame() -> pd.DataFrame:
    """pvlib 결과를 흉내낸 프레임. 인덱스도 컬럼도 tz-aware 다."""
    index = pd.date_range("2023-07-03 00:00", periods=4, freq="15min", tz="Asia/Seoul")
    return pd.DataFrame(
        {
            "생산시각": index,
            "pv_kw": [0.0, 12.5, 30.0, 41.2],
        },
        index=index,
    )


def test_strip_timezone_clears_index_and_columns() -> None:
    stripped = strip_timezone(_tz_frame())
    assert pd.DatetimeIndex(stripped.index).tz is None
    assert stripped["생산시각"].dt.tz is None
    assert stripped["생산시각"].iloc[1] == pd.Timestamp("2023-07-03 00:15")


def test_strip_timezone_does_not_mutate_the_input() -> None:
    frame = _tz_frame()
    strip_timezone(frame)
    assert pd.DatetimeIndex(frame.index).tz is not None


def test_tz_aware_frame_saves_without_error(tmp_path: Path) -> None:
    """**tz 를 해제하지 않으면 openpyxl 이 ValueError 를 낸다.** 저장이 통과해야 한다."""
    path = write_workbook({"감도": _tz_frame()}, tmp_path / "tz.xlsx")
    restored = pd.read_excel(path, sheet_name="감도", index_col=0)
    assert len(restored) == 4
    assert pd.DatetimeIndex(restored.index).tz is None


def test_period_index_becomes_text(tmp_path: Path) -> None:
    """``BillingResult.monthly`` 는 Period 인덱스다. Excel 이 쓰지 못한다."""
    frame = pd.DataFrame({"won": [1.0, 2.0]}, index=pd.period_range("2023-07", periods=2, freq="M"))
    stripped = strip_timezone(frame)
    assert list(stripped.index) == ["2023-07", "2023-08"]
    assert write_workbook({"월별 집계": frame}, tmp_path / "period.xlsx").is_file()


# --------------------------------------------------------------------- 파일명·쓰기 실패


def test_result_path_carries_a_timestamp_suffix() -> None:
    """접미사가 없으면 Excel 이 파일을 열고 있을 때 덮어쓰기가 실패한다."""
    path = result_path(Path("output"), now=dt.datetime(2026, 8, 7, 14, 30))
    assert path.name == "result_20260807_1430.xlsx"
    assert path.parent == Path("output")


def test_permission_error_tells_the_user_to_close_excel(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Excel 이 파일을 잡고 있으면 PermissionError 가 난다. 그대로 던지지 않는다."""

    def deny(*args: object, **kwargs: object) -> None:
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr("kwise.report.excel.pd.ExcelWriter", deny)
    with pytest.raises(ReportWriteError, match="Excel 에서 파일을 닫아 주세요"):
        write_workbook({"요약": pd.DataFrame({"a": [1]})}, tmp_path / "locked.xlsx")


def test_output_directory_is_created(
    sample_usage: UsageData, sample_bill: BillingResult, tmp_path: Path
) -> None:
    """``output\\`` 이 없어도 만든다. 한글 경로도 통과해야 한다."""
    sections = ReportSections(usage=sample_usage, bill=sample_bill, include_timeseries=False)
    assert export_report(sections, output_dir=tmp_path / "산출물" / "2026").is_file()


# --------------------------------------------------------------------- 요약 시트 필수 문구


def test_summary_carries_the_traceability_lines(
    summary_text: str, sample_bill: BillingResult
) -> None:
    """요구사항서 5.8 — 적용 요금표·계약종별·단가·시간대 구분·산출 기간."""
    for line in sample_bill.traceability():
        label, _, value = line.partition(": ")
        assert label in summary_text, label
        assert value in summary_text, line
    assert "2026-06-01 시행" in summary_text
    assert "일반용전력(을) 고압A 선택I" in summary_text


def test_summary_carries_the_not_included_notice(summary_text: str) -> None:
    """요구사항서 5.1 — 기타 요금요소는 미포함이며 실제 절감액은 더 크다."""
    assert NOT_INCLUDED_NOTICE in summary_text
    assert "기후환경요금" in summary_text
    assert "부가가치세" in summary_text


def test_summary_carries_the_contract_change_warning(summary_text: str) -> None:
    """요구사항서 9.4 — 한 번의 초과가 12개월간 적용된다."""
    assert CONTRACT_CHANGE_WARNING in summary_text
    assert "한 번의 초과가 12개월간 적용됩니다" in summary_text


def test_summary_carries_every_known_limit(summary_text: str) -> None:
    """요구사항서 부록 D — 열한 항목을 하나도 빠뜨리지 않는다.

    역률 정정(약관 제41·42·43조)으로 두 항목이, 경제성DR(전력시장운영규칙 제12장)로
    두 항목이 늘었다.
    """
    assert len(KNOWN_LIMITS) == 11
    for limit in KNOWN_LIMITS:
        assert limit in summary_text, limit
    assert any("제42조" in limit for limit in KNOWN_LIMITS)  # 30분 누적 계량
    assert any("제43조 ③" in limit for limit in KNOWN_LIMITS)  # 첫 달 예고
    assert any("순편익가격" in limit for limit in KNOWN_LIMITS)  # DR 단가 미산출
    assert any("별표28" in limit for limit in KNOWN_LIMITS)  # CBL


def test_summary_records_the_pv_judgement_population(summary_text: str) -> None:
    """어느 모집단으로 태양광 등급을 매겼는지 밝힌다 (5.2 ①)."""
    assert "태양광 판정 모집단" in summary_text
    assert "요금적용전력 대상 슬롯(중간·최대부하)" in summary_text
    assert "부록 B 의 시각 분포는 전 슬롯" in summary_text


def test_summary_records_the_missing_data_policy(summary_text: str) -> None:
    """결측은 보간하지 않는다. 그 사실이 산출물에 남아야 한다."""
    assert "미보간" in summary_text
    assert "972슬롯" in summary_text


def test_summary_carries_the_recalculation_note(summary_text: str) -> None:
    """조합 절감액은 합이 아니라는 것을 산출물에 적는다 (8장)."""
    assert "수단별 절감액의 합이 아닙니다" in summary_text
    assert "확실성 등급은 가장 낮은 구성 요소를 따릅니다" in summary_text


def test_summary_carries_the_quality_warnings(summary_text: str) -> None:
    assert "신뢰 제한" in summary_text
    assert "직전 12개월" in summary_text


# --------------------------------------------------------------------- 미산출은 빈칸이 아니다


def test_format_won_writes_a_reason_instead_of_a_blank() -> None:
    assert format_won(1_234_567.0) == "1,234,567"
    assert format_won(None) == UNPRICED_REASONS["contract"]
    assert format_won(None, reason=UNPRICED_REASONS["external_price"]).startswith("미산출")


def test_contract_row_states_the_basis_of_its_number(
    sample_usage: UsageData, sample_bill: BillingResult
) -> None:
    """하한 30% 가 확인된 종별이라 금액이 나온다. 근거를 함께 적는다 (6.4).

    이 건물은 하한 2,100 kW 가 요금적용전력 5,293 kW 에 걸리지 않아 0 원이다.
    """
    adjustment = evaluate_contract_adjustment(sample_usage, sample_bill, contract_kw=7_000.0)
    assert adjustment.saving_won == pytest.approx(0.0)
    row = measure_summary_frame(contract=adjustment).iloc[0]
    assert row["절감액(원)"] == "0"
    assert "하한 30%" in row["비고"]
    assert "하향 여지" in row["비고"]


def test_contract_row_shows_the_reason_when_the_floor_is_unknown(
    sample_usage: UsageData, sample_bill: BillingResult
) -> None:
    """하한 규정을 모르는 종별이면 빈칸이 아니라 사유를 적는다.

    없는 절감을 만들어내지 않는 것이 이 자리의 요점이다.
    """
    from dataclasses import replace

    known = evaluate_contract_adjustment(sample_usage, sample_bill, contract_kw=7_000.0)
    unknown = replace(known, saving_won=None, annual_saving_won=None)
    row = measure_summary_frame(contract=unknown).iloc[0]
    assert row["절감액(원)"] == UNPRICED_REASONS["contract"]
    assert row["12개월 환산(원)"] == UNPRICED_REASONS["contract"]
    assert row["회수기간"] == "—"


def test_sensitivity_sheet_says_why_it_is_empty() -> None:
    """태양광이 없으면 감도 시트가 빈 장이 아니라 사유로 채워진다 (9.2)."""
    frame = no_pv_sensitivity_frame()
    assert frame.iloc[0]["내용"] == NO_PV_SENSITIVITY_NOTE
    assert "PV 출력에만 적용" in NO_PV_SENSITIVITY_NOTE


# --------------------------------------------------------------------- 진단 시트


def test_diagnosis_sheet_keeps_the_two_populations_apart(
    sample_sheets: dict[str, pd.DataFrame],
    sample_diagnosis: Diagnosis,
) -> None:
    """부록 B 원값(전 슬롯)과 마스크 적용 값을 함께 싣되 라벨로 구분한다."""
    index = list(sample_sheets["진단"].index)
    raw = [name for name in index if name.endswith("시 (전 슬롯)")]
    masked = [name for name in index if name.endswith("시 (요금적용전력 대상)")]
    assert "상위 100구간 7시 (전 슬롯)" in raw  # 경부하라 대상에서는 빠진다
    assert "상위 100구간 7시 (요금적용전력 대상)" not in masked
    assert len(raw) == 11
    assert len(masked) == 10

    values = sample_sheets["진단"]["값"]
    assert values["요금적용전력 대상 슬롯"] == f"{sample_diagnosis.peak.demand_eligible_slots:,}"
    assert values["상위 100구간 주말 건수 (요금적용전력 대상)"] == "0"


# --------------------------------------------------------------------- 조합 절감액은 재계산이다


def test_sheet_saving_is_recalculated_not_summed(
    sample_sheets: dict[str, pd.DataFrame],
    sample_usage: UsageData,
    sample_report: QualityReport,
    tariff: TariffTable,
    sample_bill: BillingResult,
    sample_unit_pv: pd.Series,
) -> None:
    """조합 비교 시트의 절감액 ≠ 요금제 절감 + 태양광 단독 절감.

    태양광이 사용량을 줄이면 선택요금별 유불리가 달라진다. 시트에 실린 값이
    재계산 결과라는 것을 여기서 못 박는다 (요구사항서 8장).
    """
    frame = sample_sheets["조합 비교"]
    switch_only = float(frame.loc["선택요금 전환", "절감액(원)"])
    combined = float(frame.loc["+ 태양광 500 kWp", "절감액(원)"])

    pv_only = evaluate_combination(
        sample_usage,
        tariff,
        CombinationSpec(name="태양광만", selection=CURRENT, pv_capacity_kwp=500.0),
        baseline_bill=sample_bill,
        unit_pv_kw_per_kwp=sample_unit_pv,
        quality=sample_report,
    ).saving_won

    assert switch_only > 0
    assert pv_only > 0
    assert combined != pytest.approx(switch_only + pv_only, rel=1e-6)
    assert abs(combined - (switch_only + pv_only)) > 1_000_000  # 어림 오차가 아니다


def test_baseline_row_has_no_saving(sample_sheets: dict[str, pd.DataFrame]) -> None:
    baseline = sample_sheets["조합 비교"].loc["기준선 (현행)"]
    assert float(baseline["절감액(원)"]) == pytest.approx(0.0)
    assert float(baseline["투자비(원)"]) == pytest.approx(0.0)


# --------------------------------------------------------------------- 확실성은 최저 등급을 따른다


def test_certainty_in_the_sheet_follows_the_lowest_component(
    sample_sheets: dict[str, pd.DataFrame],
) -> None:
    """요금제만 '높음', 태양광이 끼면 '중간', ESS 가 끼면 '중간~낮음'."""
    certainty = sample_sheets["조합 비교"]["확실성"]
    assert certainty["기준선 (현행)"] == str(Certainty.HIGH)
    assert certainty["선택요금 전환"] == str(Certainty.HIGH)
    assert certainty["+ 태양광 500 kWp"] == str(Certainty.MEDIUM)
    # PV + ESS 조합이다. 더 낮은 ESS 등급을 따른다.
    assert certainty["+ ESS 목표 5,000 kW"] == str(Certainty.MEDIUM_LOW)


def test_measure_sheet_grades_each_measure_on_its_own(
    sample_measure_rows: pd.DataFrame,
) -> None:
    """수단별 결과는 조합이 아니므로 각자의 등급을 그대로 쓴다."""
    grades = sample_measure_rows["확실성"]
    assert grades.iloc[0] == str(Certainty.HIGH)  # 선택요금 전환
    assert grades.iloc[-1] == str(Certainty.MEDIUM_LOW)  # ESS


# --------------------------------------------------------------------- CLI 배치


@pytest.fixture
def two_case_yaml(tmp_path: Path) -> Path:
    """케이스 두 건. 합성 한 달치라 네트워크도 태양광도 타지 않는다."""
    write_month(tmp_path / "건물A.csv", 2024, 3, kwh=100.0)
    write_month(tmp_path / "건물B.csv", 2024, 3, kwh=250.0)
    path = tmp_path / "cases.yaml"
    path.write_text(
        "output_dir: out\n"
        "cases:\n"
        "  - name: 건물A\n"
        "    usage: 건물A.csv\n"
        "    contract_type: general_b\n"
        "    voltage: high_a\n"
        "    option: I\n"
        "    contract_kw: 500\n"
        "  - name: 건물B\n"
        "    usage: 건물B.csv\n"
        "    contract_type: general_b\n"
        "    voltage: high_b\n"
        "    option: II\n"
        "    contract_kw: 1200\n",
        encoding="utf-8",
    )
    return path


def test_batch_config_reads_utf8_and_resolves_paths(two_case_yaml: Path) -> None:
    """cp949 로 열리면 한글 케이스명이 깨진다. 경로는 정의 파일 기준 상대다."""
    config = load_batch_config(two_case_yaml)
    assert [case.name for case in config.cases] == ["건물A", "건물B"]
    assert config.cases[0].usage == two_case_yaml.parent / "건물A.csv"
    assert config.cases[1].selection == TariffSelection("general_b", "high_b", "II")
    assert config.output_dir == two_case_yaml.parent / "out"


def test_cli_runs_two_cases_to_completion(
    two_case_yaml: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """``python -m kwise.cli run --cases cases.yaml`` 이 끝까지 돈다."""
    monkeypatch.setenv("PROJECT_CACHE", str(tmp_path / "cache"))
    result = CliRunner().invoke(app, ["run", "--cases", str(two_case_yaml), "--no-timeseries"])
    assert result.exit_code == 0, result.output
    assert "2개 케이스" in result.output
    assert "건물A" in result.output
    assert "건물B" in result.output

    out = two_case_yaml.parent / "out"
    assert len(list(out.glob("result_건물A_*.xlsx"))) == 1
    assert len(list(out.glob("result_건물B_*.xlsx"))) == 1
    summaries = list(out.glob("summary_*.csv"))
    assert len(summaries) == 1
    # Excel 에서 열 CSV 다. utf-8-sig 로 써야 한글이 깨지지 않는다.
    assert summaries[0].read_bytes().startswith(b"\xef\xbb\xbf")
    frame = pd.read_csv(summaries[0], encoding="utf-8-sig", index_col=0)
    assert list(frame.index) == ["건물A", "건물B"]
    assert (frame["baseline_won"] > 0).all()


def test_cli_version_command() -> None:
    result = CliRunner().invoke(app, ["version"])
    assert result.exit_code == 0
    assert "kWise" in result.output


def test_resume_skips_cases_that_already_finished(
    two_case_yaml: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """중간 결과가 남아 있으면 다시 계산하지 않는다."""
    monkeypatch.setenv("PROJECT_CACHE", str(tmp_path / "cache"))
    config = load_batch_config(two_case_yaml)
    first = run_batch(config, include_timeseries=False)
    assert first.skipped == ()

    second = run_batch(config, resume=True, include_timeseries=False)
    assert second.skipped == ("건물A", "건물B")
    assert [item.name for item in second.summaries] == ["건물A", "건물B"]
    assert second.summaries[0].baseline_won == first.summaries[0].baseline_won


def test_batch_case_report_has_the_required_sheets(
    two_case_yaml: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """케이스별 통합문서에도 필수 문구가 실린다. 시계열만 빠진다."""
    monkeypatch.setenv("PROJECT_CACHE", str(tmp_path / "cache"))
    run_batch(load_batch_config(two_case_yaml), include_timeseries=False)
    path = next((two_case_yaml.parent / "out").glob("result_건물A_*.xlsx"))
    workbook = load_workbook(path, read_only=True)
    try:
        assert "15분 시계열" not in workbook.sheetnames
        assert workbook.sheetnames == [name for name in SHEET_ORDER if name != "15분 시계열"]
    finally:
        workbook.close()
    summary = pd.read_excel(path, sheet_name="요약", index_col=0)
    text = "\n".join(str(value) for value in summary.to_numpy().ravel())
    assert CONTRACT_CHANGE_WARNING in text
    assert NOT_INCLUDED_NOTICE in text
    assert KNOWN_LIMITS[0] in text
    assert NO_PV_SENSITIVITY_NOTE in pd.read_excel(path, sheet_name="감도").to_string()


def test_empty_case_list_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "empty.yaml"
    path.write_text("cases: []\n", encoding="utf-8")
    with pytest.raises(ValueError, match="케이스 정의가 비어 있습니다"):
        load_batch_config(path)
